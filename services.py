import os
import time
import zipfile
import asyncio
import threading
from typing import Dict, Optional
from datetime import datetime

from config import Config
from models import DownloadTask
from downloader.twitter_downloader import TwitterDownloader
from logger import DownloadLogger
from realtime_logger import log_manager
import database


class DownloadService:
    """下载服务"""
    
    def __init__(self):
        self._tasks: Dict[str, DownloadTask] = {}
        self._lock = threading.Lock()
        
        # 确保下载目录存在
        os.makedirs(Config.DOWNLOAD_FOLDER, exist_ok=True)
        
        # 确保日志目录存在
        os.makedirs(DownloadLogger.LOG_DIR, exist_ok=True)
        
        # 初始化数据库
        database.init_db()
    
    def get_task(self, task_id: str) -> Optional[DownloadTask]:
        """获取任务（优先内存，其次数据库）"""
        # 先从内存获取
        task = self._tasks.get(task_id)
        if task:
            return task
        
        # 内存中没有，从数据库获取
        history = database.get_download_by_task_id(task_id)
        if history:
            task = DownloadTask(task_id, history['user_id'], account_user_id=history.get('account_user_id'))
            task.status = history['status']
            task.total_files = history.get('total_files', 0)
            task.downloaded_files = history.get('downloaded_files', 0)
            task.zip_path = history.get('zip_path')
            task.error_message = history.get('error_message')
            task.progress = 100 if history['status'] == 'completed' else 0
            return task
        
        return None
    
    def create_task(self, user_id: str, download_type: str = 'all', account_user_id: int = None, export_xlsx: bool = False, create_zip: bool = True) -> DownloadTask:
        """创建下载任务"""
        # 检查是否已有该用户的下载记录
        existing_record = database.get_latest_download_by_user_id(user_id)
        
        if existing_record:
            # 复用之前的task_id
            task_id = existing_record['task_id']
            # 重置状态为downloading
            database.update_download_history(
                task_id,
                status='downloading',
                downloaded_files=0,
                total_files=0,
                error_message=None,
                completed_at=None
            )
        else:
            # 创建新的task_id
            task_id = f"{user_id}_{int(time.time())}"
            # 添加到数据库
            database.add_download_history(task_id, user_id, account_user_id=account_user_id)
        
        task = DownloadTask(task_id, user_id, download_type, account_user_id=account_user_id, export_xlsx=export_xlsx, create_zip=create_zip)
        
        with self._lock:
            self._tasks[task_id] = task
        
        # 记录实时日志
        log_manager.info(task_id, user_id, f'下载任务已创建: {user_id}', 'system')
        
        # 在后台线程中启动下载
        thread = threading.Thread(target=self._run_download, args=(task_id,))
        thread.daemon = True
        thread.start()
        
        return task
    
    def _run_download(self, task_id: str):
        """运行下载任务（在后台线程中执行）"""
        task = self.get_task(task_id)
        if not task:
            return
        
        # 初始化日志器
        file_logger = DownloadLogger(task_id, task.user_id)
        
        try:
            task.status = 'downloading'
            database.update_download_history(task_id, status='downloading')
            log_manager.info(task_id, task.user_id, '开始下载...', 'system')
            
            # 创建用户下载目录
            user_download_path = os.path.join(Config.DOWNLOAD_FOLDER, task.user_id)
            os.makedirs(user_download_path, exist_ok=True)
            task.download_path = user_download_path
            log_manager.info(task_id, task.user_id, f'下载目录: {user_download_path}', 'system')
            
            # 进度回调函数
            def progress_callback(progress: int, downloaded_files: int, total_files: int, skipped_files: int = 0):
                task.progress = progress
                task.downloaded_files = downloaded_files
                task.total_files = total_files
                # 更新数据库
                database.update_download_history(
                    task_id,
                    downloaded_files=downloaded_files,
                    total_files=total_files
                )
            
            # 用户信息回调函数
            def user_info_callback(user_name: str, avatar_url: str):
                database.update_download_history(
                    task_id,
                    user_name=user_name,
                    avatar_url=avatar_url
                )
            
            # 从数据库获取配置（使用用户特定的配置）
            proxy = Config.get_proxy(user_id=task.account_user_id)
            cookie = Config.get_cookie(user_id=task.account_user_id)
            
            # 创建下载器
            downloader = TwitterDownloader(
                user_id=task.user_id,
                download_path=user_download_path,
                proxy=proxy,
                cookie=cookie,
                task_id=task_id,
                progress_callback=progress_callback,
                user_info_callback=user_info_callback,
                skip_existing=True,
                max_retries=50
            )
            
            # 将实时日志管理器传递给下载器
            downloader.set_log_manager(log_manager)
            
            # 创建新的事件循环来运行异步下载
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            try:
                result = loop.run_until_complete(downloader.start_download())
                downloaded = result.get('downloaded_files', 0)
                skipped = result.get('skipped_files', 0)
                failed = result.get('failed_files', 0)
                
                task.total_files = downloaded + skipped + failed
                task.downloaded_files = downloaded + skipped  # 跳过的也算已完成
                task.skipped_files = skipped
                task.failed_files = failed
                task.tweets_info = result.get('tweets_info', [])
                task.progress = 100
                
                log_manager.success(
                    task_id, task.user_id,
                    f'下载完成 - 新增: {downloaded}, 跳过: {skipped}, 失败: {failed}',
                    'system'
                )
                
                # 更新数据库
                database.update_download_history(
                    task_id,
                    user_name=result.get('user_name'),
                    avatar_url=result.get('avatar_url'),
                    total_files=task.total_files,
                    downloaded_files=task.downloaded_files
                )
            finally:
                loop.close()
            
            # 创建ZIP文件（如果启用）
            if task.create_zip:
                log_manager.info(task_id, task.user_id, '正在创建ZIP压缩文件...', 'system')
                self._create_zip(task)
                log_manager.success(task_id, task.user_id, f'ZIP文件创建完成: {os.path.basename(task.zip_path)}', 'system')
            else:
                log_manager.info(task_id, task.user_id, '跳过创建ZIP压缩包', 'system')
            
            task.status = 'completed'
            task.end_time = time.time()
            
            # 计算ZIP文件大小
            file_size = 0
            if task.zip_path and os.path.exists(task.zip_path):
                file_size = os.path.getsize(task.zip_path)
            
            # 更新数据库
            database.update_download_history(
                task_id,
                status='completed',
                zip_path=task.zip_path,
                file_size=file_size,
                completed_at=time.strftime('%Y-%m-%d %H:%M:%S')
            )
            
            log_manager.success(task_id, task.user_id, '任务已完成!', 'system')
            
        except Exception as e:
            task.status = 'failed'
            task.error_message = str(e)
            task.end_time = time.time()
            
            log_manager.error(task_id, task.user_id, f'任务失败: {str(e)}', 'system')
            
            # 更新数据库
            database.update_download_history(
                task_id,
                status='failed',
                error_message=str(e),
                completed_at=time.strftime('%Y-%m-%d %H:%M:%S')
            )
        
        finally:
            # 清理内存中的任务（延迟清理，保持一段时间可查询）
            def cleanup():
                with self._lock:
                    if task_id in self._tasks:
                        del self._tasks[task_id]
            timer = threading.Timer(3600, cleanup)  # 1小时后清理
            timer.daemon = True
            timer.start()
    
    def clear_all_cache(self) -> dict:
        """清理所有缓存文件和下载历史"""
        import shutil
        
        deleted_files = 0
        deleted_dirs = 0
        
        download_folder = Config.DOWNLOAD_FOLDER
        
        if os.path.exists(download_folder):
            try:
                # 遍历下载目录
                for item in os.listdir(download_folder):
                    item_path = os.path.join(download_folder, item)
                    
                    # 跳过隐藏文件
                    if item.startswith('.'):
                        continue
                    
                    if os.path.isfile(item_path):
                        # 删除文件（ZIP文件等）
                        os.remove(item_path)
                        deleted_files += 1
                    elif os.path.isdir(item_path):
                        # 删除用户下载目录
                        file_count = sum(len(files) for _, _, files in os.walk(item_path))
                        shutil.rmtree(item_path)
                        deleted_dirs += 1
                        deleted_files += file_count
            except Exception as e:
                raise Exception(f'清理缓存失败: {str(e)}')
        
        # 清空下载历史
        database.clear_all_download_history()
        
        return {'deleted_files': deleted_files, 'deleted_dirs': deleted_dirs}
    
    def _create_zip(self, task: DownloadTask):
        """创建ZIP压缩文件，按文件类型分类"""
        # 文件类型分类
        video_exts = {'.mp4', '.mov', '.avi', '.mkv', '.webm', '.gif'}
        image_exts = {'.jpg', '.jpeg', '.png', '.webp', '.bmp'}
        
        def get_file_type(filename: str) -> str:
            """根据文件扩展名返回文件类型"""
            ext = os.path.splitext(filename)[1].lower()
            if ext in video_exts:
                return 'video'
            elif ext in image_exts:
                return 'image'
            else:
                return 'other'
        
        def get_category(filename: str) -> str:
            """根据文件扩展名返回分类文件夹名"""
            file_type = get_file_type(filename)
            if file_type == 'video':
                return 'videos'
            elif file_type == 'image':
                return 'images'
            else:
                return 'others'
        
        # 生成时间戳
        timestamp = time.strftime('%Y%m%d_%H%M%S')
        
        # 获取用户信息
        from database import get_download_by_task_id
        history = get_download_by_task_id(task.task_id)
        user_name = history.get('user_name') if history else None
        
        # 构建名称：用户名_用户ID（如果有用户名）
        name_prefix = f'{user_name}_{task.user_id}' if user_name else task.user_id
        
        # 根据下载类型生成文件名
        download_type = task.download_type
        if download_type == 'video':
            zip_filename = f'{name_prefix}_video_{timestamp}.zip'
        elif download_type == 'image':
            zip_filename = f'{name_prefix}_img_{timestamp}.zip'
        else:  # all
            zip_filename = f'{name_prefix}_video_img_{timestamp}.zip'
        
        zip_path = os.path.join(Config.DOWNLOAD_FOLDER, zip_filename)
        
        # 构建根目录名称
        folder_name = name_prefix
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(task.download_path):
                for file in files:
                    file_type = get_file_type(file)
                    
                    # 根据下载类型过滤文件
                    if download_type == 'video' and file_type != 'video':
                        continue
                    if download_type == 'image' and file_type != 'image':
                        continue
                    
                    file_path = os.path.join(root, file)
                    # 获取分类文件夹
                    category = get_category(file)
                    # 构建ZIP内的路径：根目录/分类文件夹/原文件名
                    arcname = f'{folder_name}/{category}/{file}'
                    zipf.write(file_path, arcname)
            
            # 如果需要导出xlsx
            if task.export_xlsx and task.tweets_info:
                log_manager.info(task.task_id, task.user_id, '正在生成xlsx文件...', 'system')
                xlsx_path = self._create_xlsx(task)
                if xlsx_path:
                    xlsx_filename = f'{task.user_id}_tweets.xlsx'
                    zipf.write(xlsx_path, f'{folder_name}/{xlsx_filename}')
                    log_manager.success(task.task_id, task.user_id, f'xlsx文件生成完成，共 {len(task.tweets_info)} 条记录', 'system')
                    # 清理临时xlsx文件
                    try:
                        os.remove(xlsx_path)
                    except Exception:
                        pass
                else:
                    log_manager.error(task.task_id, task.user_id, 'xlsx文件生成失败', 'system')
        
        task.zip_path = zip_path
    
    def _create_xlsx(self, task: DownloadTask) -> str:
        """创建xlsx文件，返回文件路径"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = '推文数据'
            
            # 定义表头
            headers = ['发布时间', '用户名', '用户ID', '推文内容', '媒体类型', '媒体URL', '下载URL']
            
            # 表头样式
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4A6CF7', end_color='4A6CF7', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 写入数据
            for row_idx, tweet in enumerate(task.tweets_info, 2):
                ws.cell(row=row_idx, column=1, value=tweet.get('time', '')).border = thin_border
                ws.cell(row=row_idx, column=2, value=tweet.get('name', '')).border = thin_border
                ws.cell(row=row_idx, column=3, value=tweet.get('screen_name', '')).border = thin_border
                
                # 推文内容，限制长度避免单元格过大
                text = tweet.get('text', '')
                if len(text) > 500:
                    text = text[:500] + '...'
                cell = ws.cell(row=row_idx, column=4, value=text)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                ws.cell(row=row_idx, column=5, value=tweet.get('type', '')).border = thin_border
                ws.cell(row=row_idx, column=6, value=tweet.get('media_url', '')).border = thin_border
                ws.cell(row=row_idx, column=7, value=tweet.get('download_url', '')).border = thin_border
            
            # 设置列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 60
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 40
            ws.column_dimensions['G'].width = 40
            
            # 保存到临时文件
            xlsx_path = os.path.join(Config.DOWNLOAD_FOLDER, f'{task.user_id}_{task.task_id}_tweets.xlsx')
            wb.save(xlsx_path)
            return xlsx_path
            
        except Exception as e:
            error_msg = f'创建xlsx失败: {e}'
            print(error_msg)
            log_manager.error(task.task_id, task.user_id, error_msg, 'system')
            return None


# 全局下载服务实例
download_service = DownloadService()
